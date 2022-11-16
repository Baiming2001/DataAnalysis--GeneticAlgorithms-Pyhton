import openpyxl,os
eternal_path= os.path.join('eternal.xlsx')
eb=openpyxl.load_workbook(eternal_path)
es=eb["Sheet1"]
s1=[13, 12, 14, 4, 3, 2, 5, 8, 1, 6, 7, 10, 11, 9]
s2=[13, 12, 4, 3, 2, 5, 8, 1, 6, 7, 15, 10, 11, 9]
s3=[13, 12, 4, 16, 3, 2, 5, 8, 1, 6, 7, 10, 11, 9]
s4=[13, 2, 12, 4, 3, 5, 8, 1, 17, 6, 7, 10, 11, 9]
s5=[7, 6, 8, 11, 10, 15]
s6=[9, 13, 12, 14, 3, 2, 5]
s7=[16,1,4]
"""
workbook_path= os.path.join('1.xlsx')
wb=openpyxl.load_workbook(workbook_path)
ws=wb["Sheet1"]
wa = wb.create_sheet('-')
for row in ws.iter_rows():
    line = [cell.value for cell in row]
    if'(51, 14)'in line and'(48, 16)' in line:
        wa.append(line)
wb.save('sh.xlsx')
print('ok')
"""
def classify(s,num):
    workbook_path= os.path.join('1.xlsx')
    wb=openpyxl.load_workbook(workbook_path)
    ws=wb["Sheet1"]
    weg= os.path.join('1.xlsx')
    gb=openpyxl.load_workbook(weg)
    length=len(s)-1
    num=str(num)
    for i in range(0,length):
            a=es.cell(row = s[i], column = 5).value
            b=es.cell(row = s[i+1], column = 5).value
            c=str(i)
            ga = gb.create_sheet(c)
            for row in ws.iter_rows():
                line = [cell.value for cell in row]
                if a in line and b in line:
                    ga.append(line)
            print('+'+num)
    a=es.cell(row = s[length], column = 5).value
    b=es.cell(row = s[0], column = 5).value
    c=str(i+1)
    ga = gb.create_sheet(c)
    for row in ws.iter_rows():
            line = [cell.value for cell in row]
    if a in line and b in line:
            ga.append(line)
    print(num+'finisched')
    gb.save(num+'Haupt.xlsx')
    wb.save(workbook_path)
    print('one finisched')
    for k in range(2,3):
        k1=str(k)
        workbook_path= os.path.join(k1+'.xlsx')
        wb=openpyxl.load_workbook(workbook_path)
        ws=wb["Sheet1"]
        for i in range(0,length):
            a=es.cell(row = s1[i], column = 5).value
            b=es.cell(row = s1[i+1], column = 5).value
            c=str(i)
            ga = gb.get_sheet_by_name(c)
            for row in ws.iter_rows():
                line = [cell.value for cell in row]
                if a in line and b in line:
                    ga.append(line)
            print('+'+num)
        a=es.cell(row = s[length], column = 5).value
        b=es.cell(row = s[0], column = 5).value
        c=str(i+1)
        ga = gb.get_sheet_by_name(c)
        for row in ws.iter_rows():
                line = [cell.value for cell in row]
                if a in line and b in line:
                    ga.append(line)
        print(c)
        gb.save(num+'Haupt.xlsx')
        wb.save(workbook_path)
        print(k1+'finisched')
    print(num+"ok")
#执行
classify(s1,1)
classify(s2,2)
classify(s3,3)
classify(s4,4)
classify(s5,5)
classify(s6,6)
classify(s7,7)
#wb.remove_sheet(ws)

"""
workbook_path= os.path.join('sh.xlsx')
wb=openpyxl.load_workbook(workbook_path)
ws=wb["Sheet1"]
wa=wb["-6-3"]
for row in ws.iter_rows():
    line = [cell.value for cell in row]
    if '(54, 14)' in line and '(47, 16)' in line:
        wa.append(line)
    print('+2')
wb.save('sh.xlsx')
print('ok')
"""

